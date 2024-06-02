import openpyxl
import sqlite3


path_bd = 'db.sqlite'
con = sqlite3.connect(path_bd)
cur = con.cursor()
index = 0


def create_bd():
    cur.execute('''
      CREATE TABLE IF NOT EXISTS goods(
        id INTEGER PRIMARY KEY,
        Ccode TEXT,
        Vcode TEXT,
        art TEXT,
        part TEXT,
        price_inside INTEGER,
        inv TEXT,
        price_outside INTEGER
        );
    ''')


def import_table_to_bd(path_intake):
    print(path_intake)
    cur = con.cursor()
    wb = openpyxl.load_workbook(filename=path_intake, read_only=True)
    list_name = wb.sheetnames
    sheet = wb[list_name[0]]
    i = index
    num = len(tuple(sheet.rows))
    for row in sheet.iter_rows(min_row=1,
                               max_row=num,
                               min_col=1,
                               max_col=7,
                               values_only=True):
        i = i + 1
        line = (i, row[0], row[1], row[2], row[3], row[4], row[5], row[6])
        print(line)
        cur.execute(
            'INSERT INTO goods VALUES(?, ?, ?, ?, ?, ?, ?, ?);',
            line
        )


def main():
    path_intake = input("Введите ссылку на exel документ:")
    create_bd()
    import_table_to_bd(path_intake)
    con.commit()
    con.close()
    return 'Выполнено! База данных обновлена!'


if __name__ == '__main__':
    main()
